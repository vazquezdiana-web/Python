from openpyxl import load_workbook
wb = load_workbook("ventas_semanales.xlsx")
ws = wb["Ventas"]
ws["G1"] = "Total"

for fila in range(2, ws.max_row + 1):
    cantidad = ws[f"E{fila}"].value
    precio = ws[f"F{fila}"].value
    ws[f"G{fila}"] = cantidad * precio
    total_general = sum(ws[f"G{fila}"].value for fila in range(2, ws.max_row+ 1))
    ws[f"F{ws.max_row + 2}"] = "Total General"
    ws[f"G{ws.max_row + 2}"] = total_general
    wb.save("reporte_ventas.xlsx")

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

def generar_pdf():
    c = canvas.Canvas("reporte_ventas.pdf", pagesize=letter)
    c.setFont("Helvetica", 12)
    c.drawString(50, 750, "Reporte Semanal de Ventas")
    y = 720
    c.drawString(50, y, "ID Vendedor Región Producto Cantidad Precio Total")
    y -= 20

    for fila in ws.iter_rows(min_row=2, max_row=6, values_only=True):
        total = fila[4] * fila[5]
        linea = f"{fila[0]} {fila[1]} {fila[2]} {fila[3]} {fila[4]} {fila[5]} {total}"
        c.drawString(50, y, linea)
        y -= 20
    c.drawString(50, y - 10, f"Total general: {total_general}")
    c.save()
generar_pdf()